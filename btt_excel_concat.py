import re
import time
import tkinter
import warnings
from datetime import datetime
from tkinter import filedialog
import xlwings as xw

import pandas as pd
import xxhash
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import Rule
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList


def get_max_row(sheet, start_col, end_col):
	"""
	Parameters
	----------
	:param sheet:
		Worksheet to be searched
	:param start_col:
		starting column of the column-range
	:param end_col:
		end column of the column-range
	:return:
		the last non-empty row in given column-range
	"""
	max_row = 0
	for col in range(start_col, end_col + 1):
		for row in range(1, sheet.max_row + 1):
			if sheet.cell(row=row, column=col).value is not None:
				max_row = max(max_row, row)
	return max_row


def copy_range(start_col, start_row, end_col, end_row, sheet):
	"""
	Notes
	-----
	:requirement 04.06.2024 BWB:
		 the funktion only copies the values if the flag in column "AT" is not set

	Parameters
	----------
	:param start_col:
		start column copy area
	:param start_row:
		start row of copy area
	:param end_col:
		end column of copy area
	:param end_row:
		end row of copy area
	:param sheet:
		Worksheet to be copied from
	:return:
		list of cell values in given range
	"""
	return [[sheet.cell(row=i, column=j).value for j in range(start_col, end_col + 1)] for i in range(start_row, end_row + 1) if sheet.cell(row=i, column=column_index_from_string("AT")).value != "x"]


def paste_range(start_col, start_row, sheet_receiving, copied_data):
	"""
	Parameters
	----------
	:param start_col:
		start column copy area
	:param start_row:
		start row of copy area
	:param sheet_receiving:
		receiving Worksheet to write copied data to
	:param copied_data:
		list of cell values, procured by the copy_range()-function
	:return:
		None
	"""
	for i, row_data in enumerate(copied_data, start=start_row):
		for j, value in enumerate(row_data, start=start_col):
			sheet_receiving.cell(row=i, column=j).value = value


def add_dv(sheet, dv_list):
	""" Function to add data validation to a given Worksheet
		The ranges and source values as well as the target column(s) are given by the dv_list parameter.
		Every column for each formula is assigned a data validation.

	Parameters
	----------
	:param sheet:
		Worksheet to which data validation should be added
	:param dv_list:
		list of data validations, consisting of a source table selection and one or more target columns
	:return:
		None
	"""
	range_str = ""
	max_row = sheet.max_row
	for formula, cols in dv_list.items():
		for col in cols:
			col_letter = col if col.isalpha() else get_column_letter(col)
			dv = DataValidation(
				type="list",
				formula1=f"{formula}",
				allow_blank=True,
				showDropDown=False,
				showInputMessage=True,
				showErrorMessage=True
			)
			if sheet.title == "BTT":
				range_str = f"${col_letter}$3:${col_letter}${max_row}"
			elif sheet.title == "Übersicht":
				range_str = f"A1"
			dv.add(range_str)
			sheet.add_data_validation(dv)


def add_cf(sheet, cf_list):
	""" Function to add conditional formatting to a given Worksheet.
		The ranges and formulas as well as the target column(s) are given by the cf_list parameter.
		Every column for each formula is assigned a formatting rule with a red PatternFill.

	Parameters
	----------
	:param sheet:
		Worksheet to which conditional formatting should be added
	:param cf_list:
		list of conditional formatting, consisting of a formula and one or more target columns
	:return:
		None
	"""
	red_fill = PatternFill(patternType=None,
						   fgColor=Color(rgb="000000",
										 type="rgb"),
						   bgColor=Color(rgb="FFA7A7",
										 type="rgb")
						   )
	dxf = DifferentialStyle(fill=red_fill)
	for formula, cols in cf_list.items():
		for col in cols:
			col_letter = col if col.isalpha() else get_column_letter(col)
			sheet.conditional_formatting.add(f"{col_letter}3:{col_letter}{sheet.max_row}",
											 Rule(type="expression",
												  dxf=dxf,
												  formula=[f"{formula.replace('~', col_letter)}"])
											 )


def hash_row(row):
	"""
	Parameters
	----------
	:param row:
		row to be hashed
	:return:
		hash value of the input row
	"""
	row_str = ''.join([str(cell) for cell in row])
	return xxhash.xxh64(row_str).hexdigest()


def clean_table(sheet, table_name):
	""" Function to clear a table in a given Worksheet.
		Every empty row is being eliminated by shifting non-empty rows to the top.

		The function also consumes the update_table_dimensions()-function to further apply the new dimensions to the table object after the shifting.

	Parameters
	----------
	:param sheet:
		Worksheet, that contains the table object
	:param table_name:
		table name of the table to be cleaned up
	:return:
		None
	"""
	if table_name in sheet.tables:
		table = sheet.tables[table_name]
		min_col, min_row, max_col, max_row = range_boundaries(table.ref)
		non_empty_rows = []

		for row in range(min_row, max_row + 1):
			if not all(sheet.cell(row=row, column=col).value is None for col in range(min_col, max_col + 1)):
				non_empty_rows.append(row)

		for i, row in enumerate(non_empty_rows, start=min_row):
			for col in range(min_col, max_col + 1):
				sheet.cell(row=i, column=col).value = sheet.cell(row=row, column=col).value
				if i != row:
					sheet.cell(row=row, column=col).value = None

		update_table_dimensions(sheet, table_name, min_col, max_col)


def update_table_dimensions(sheet, table_name, start_col, end_col):
	""" Function to update the table dimensions of a given table on a given Worksheet.
		Consumed by the clean_table()-function

	:param sheet:
		Worksheet containing the table to be updated
	:param table_name:
		name of thr table to be updated
	:param start_col:
		start column of the table, appointed within the clean_table()-function by taking the range boundaries of the table
	:param end_col:
		end column of the table, appointed within the clean_table()-function by taking the range boundaries of the table
	:return:
		None
	"""
	if table_name in sheet.tables:
		table = sheet.tables[table_name]
		start_row = 1 if table_name != "BTT" else 2
		table.ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{get_max_row(sheet, start_col, end_col)}"


def btt_concat(split=None):
	""" Function for merging multipie BTT-Files.

		Loads the template file and opens a file dialog to select BTT-Files to be consolidated.
		Deletes the Worksheet "Quercheck Transaktionen".
		For each file and each Worksheet within, a dictionary is built to store the dimensions of all subareas.

		Then for each file in the dictionary, unique rows are hashed and mapped to the seen_hashes dictionary.
		If unique, the values of that row get pasted to the template by consuming the paste_range()-function.

		Afterward, the tables for each Worksheet in each file get cleaned up, data validation and conditional formatting will be reset and redistributed.

		At the end, the modified template is written to an output file.

	Parameters
	----------
	:return:
		None
	"""
	root = tkinter.Tk()
	root.withdraw()
	wb = load_workbook('BTT_Template.xlsx')
	btt = wb["BTT"]
	file_paths = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx *.xls")])
	all_tab_data = {}

	tp_list = set()
	wb_list = {}

	for path in file_paths:
		wb_tmp = load_workbook(path)
		try:
			wb_tmp.remove(wb_tmp["Quercheck Transaktionen"])
		except Exception as e:
			print(str(e))

		tab_data = {}
		for sheet_name in wb_tmp.sheetnames:
			if sheet_name == "Übersicht":
				tab_data[sheet_name] = [
					{"start_col": 1, "start_row": 4, "end_col": 2, "end_row": get_max_row(wb_tmp[sheet_name], 1, 2)},  # [A:B]
					{"start_col": 5, "start_row": 2, "end_col": 8, "end_row": get_max_row(wb_tmp[sheet_name], 5, 8)}  # [E:H]
				]
			elif sheet_name == "BTT":
				tab_data[sheet_name] = [
					{"start_col": 1, "start_row": 3, "end_col": wb_tmp[sheet_name].max_column, "end_row": wb_tmp[sheet_name].max_row}
				]

			elif sheet_name == "BPML":
				tab_data[sheet_name] = [
					{"start_col": 1, "start_row": 2, "end_col": 4, "end_row": get_max_row(wb_tmp[sheet_name], 1, 4)},  # [A:D]
					{"start_col": 6, "start_row": 2, "end_col": 10, "end_row": get_max_row(wb_tmp[sheet_name], 6, 10)}  # [F:J]
				]
			elif sheet_name == "Transaktionen":
				tab_data[sheet_name] = [
					{"start_col": 1, "start_row": 2, "end_col": 7, "end_row": get_max_row(wb_tmp[sheet_name], 1, 7)},  # [A:G]
				]
			elif sheet_name == "Formulare":
				tab_data[sheet_name] = [
					{"start_col": 1, "start_row": 2, "end_col": 3, "end_row": get_max_row(wb_tmp[sheet_name], 1, 3)},  # [A:G]
				]
			elif sheet_name == "Schnittstellen":
				tab_data[sheet_name] = [
					{"start_col": 1, "start_row": 2, "end_col": 6, "end_row": get_max_row(wb_tmp[sheet_name], 1, 6)},  # [A:F]
					{"start_col": 8, "start_row": 2, "end_col": 10, "end_row": get_max_row(wb_tmp[sheet_name], 8, 10)}  # [F:J]
				]
			elif sheet_name == "Datengrundlage adesso":
				tab_data[sheet_name] = [
					{"start_col": 1, "start_row": 2, "end_col": 3, "end_row": get_max_row(wb_tmp[sheet_name], 1, 3)},  # [A:C]
					{"start_col": 5, "start_row": 2, "end_col": 5, "end_row": get_max_row(wb_tmp[sheet_name], 5, 5)},  # [E]
					{"start_col": 7, "start_row": 2, "end_col": 7, "end_row": get_max_row(wb_tmp[sheet_name], 7, 7)},  # [G]
					{"start_col": 9, "start_row": 2, "end_col": 9, "end_row": get_max_row(wb_tmp[sheet_name], 9, 9)},  # [I]
					{"start_col": 11, "start_row": 2, "end_col": 11, "end_row": get_max_row(wb_tmp[sheet_name], 11, 11)}  # [K]
				]
			else:
				tab_data[sheet_name] = {
					"start_col": wb_tmp[sheet_name].min_column,
					"start_row": wb_tmp[sheet_name].min_row,
					"end_col": wb_tmp[sheet_name].max_column,
					"end_row": wb_tmp[sheet_name].max_row
				}
		all_tab_data[path] = tab_data

	if split:
		for tp in tp_list:
			wb_list[tp] = load_workbook('BTT_Template.xlsx')

	seen_hashes = {}
	project_data = {}
	for file, data in all_tab_data.items():
		wb_tmp = load_workbook(file)
		active_tp = tp_mapping.get(wb_tmp["Übersicht"]["A1"].value)
		for sheet_name, dim in data.items():
			if sheet_name == "BTT":
				
				# TODO
				#	- works, but now more cases need to be handled
				#		- if there are values in column AS, they should be prioritised and put in column A instead							#DONE
				#		- also this behaviour should only occur when the file is NOT split													#DONE
				#		- if split, then instead of existing values, the formula should be kept in column A
				#		
				#		- for column E there is also a special case:
				#			- if there is a value in column F, this value should be pasted in colum E of that row ('manual change')			#DONE
				if not split:
					df_a = pd.read_excel(file, sheet_name=sheet_name, engine="openpyxl", skiprows=1, usecols=[0])
					df_as = pd.read_excel(file, sheet_name=sheet_name, engine="openpyxl", skiprows=1, usecols=[44])

					df_e = pd.read_excel(file, sheet_name=sheet_name, engine="openpyxl", skiprows=1, usecols=[4])
					df_f = pd.read_excel(file, sheet_name=sheet_name, engine="openpyxl", skiprows=1, usecols=[5])

					for idx in range(len(df_a)):
						if pd.isnull(df_as.iloc[idx, 0]):
							wb_tmp[sheet_name].cell(row=idx+3, column=1).value = df_a.iloc[idx, 0] 
							# TODO -> works, should always be applied tho
							wb_tmp[sheet_name].cell(row=idx+3, column=45).value = df_a.iloc[idx, 0]
						else:
							wb_tmp[sheet_name].cell(row=idx+3, column=1).value = df_as.iloc[idx, 0]

						if pd.isnull(df_f.iloc[idx, 0]):
							wb_tmp[sheet_name].cell(row=idx+3, column=5).value = df_e.iloc[idx, 0]	
						else:
							wb_tmp[sheet_name].cell(row=idx+3, column=5).value = df_f.iloc[idx, 0]	

				if isinstance(dim, list):
					for d in dim:
						sel_range = copy_range(d["start_col"], d["start_row"], d["end_col"], d["end_row"], wb_tmp[sheet_name])
						for row in sel_range:
							row_hash = hash_row(row)
							if row_hash not in seen_hashes.get(sheet_name, set()):
								seen_hashes.setdefault(sheet_name, set()).add(row_hash)
								paste_range(d["start_col"], d["start_row"] + len(seen_hashes[sheet_name]) - 1, wb[sheet_name], [row])
				else:
					sel_range = copy_range(dim["start_col"], dim["start_row"], dim["end_col"], dim["end_row"], wb_tmp[sheet_name])
					for row in sel_range:
						row_hash = hash_row(row)
						if row_hash not in seen_hashes.get(sheet_name, set()):
							seen_hashes.setdefault(sheet_name, set()).add(row_hash)
							paste_range(dim["start_col"], dim["start_row"] + len(seen_hashes[sheet_name]) - 1, wb[sheet_name], [row])
			else:
				if isinstance(dim, list):
					for d in dim:
						sel_range = copy_range(d["start_col"], d["start_row"], d["end_col"], d["end_row"], wb_tmp[sheet_name])
						for row in sel_range:
							row_hash = hash_row(row)
							if row_hash not in seen_hashes.get(sheet_name, set()):
								seen_hashes.setdefault(sheet_name, set()).add(row_hash)
								paste_range(d["start_col"], d["start_row"] + len(seen_hashes[sheet_name]) - 1, wb[sheet_name], [row])
				else:
					sel_range = copy_range(dim["start_col"], dim["start_row"], dim["end_col"], dim["end_row"], wb_tmp[sheet_name])
					for row in sel_range:
						row_hash = hash_row(row)
						if row_hash not in seen_hashes.get(sheet_name, set()):
							seen_hashes.setdefault(sheet_name, set()).add(row_hash)
							paste_range(dim["start_col"], dim["start_row"] + len(seen_hashes[sheet_name]) - 1, wb[sheet_name], [row])

			for tab_name in wb[sheet_name].tables:
				clean_table(wb[sheet_name], tab_name)
			for tp, _wb in wb_list.items():
				for tab_name in _wb[sheet_name].tables:
					clean_table(_wb[sheet_name], tab_name)

		wb_tmp.close()

		for idx in range(3, wb["BTT"].max_row + 1):
			wb["BTT"].cell(row=idx, column=1).value = re.sub(r'aktives_Teilprojekt', f'"{active_tp}"', str(wb["BTT"].cell(row=idx, column=1).value))

	dv_list = {
		f'BPML!$A$2:$A${get_max_row(wb["BPML"], column_index_from_string("A"), column_index_from_string("A"))}': {'B'},
		f'=BPML!$F$2:$F${get_max_row(wb["BPML"], column_index_from_string("F"), column_index_from_string("F"))}': {'C'},
		f'=Transaktionen!$A$2:$A${get_max_row(wb["Transaktionen"], column_index_from_string("A"), column_index_from_string("A"))}': {'I'},
		f'=Schnittstellen!$H$2:$H${get_max_row(wb["Schnittstellen"], column_index_from_string("H"), column_index_from_string("H"))}': {'R'},
		f'=\'Datengrundlage adesso\'!$A$2:$A${get_max_row(wb["Datengrundlage adesso"], column_index_from_string("A"), column_index_from_string("A"))}': {'H'},
		f'=\'Datengrundlage adesso\'!$E$2:$E${get_max_row(wb["Datengrundlage adesso"], column_index_from_string("E"), column_index_from_string("E"))}': {'Z'},
		f'=\'Datengrundlage adesso\'!$G$2:$G${get_max_row(wb["Datengrundlage adesso"], column_index_from_string("G"), column_index_from_string("G"))}': {'X', 'AA', 'AB', 'O', 'AG', 'AH', 'AI', 'AJ'},
		f'=\'Datengrundlage adesso\'!$I$2:$I${get_max_row(wb["Datengrundlage adesso"], column_index_from_string("I"), column_index_from_string("I"))}': {'T'},
		f'=Formulare!$A$2:$A${get_max_row(wb["Formulare"], column_index_from_string("A"), column_index_from_string("A"))}': {'U'},
		f'=\'Datengrundlage adesso\'!$K$2:$K${get_max_row(wb["Datengrundlage adesso"], column_index_from_string("K"), column_index_from_string("K"))}': {'AD'},
		f'=Übersicht!$F$2:$F${get_max_row(wb["Übersicht"], column_index_from_string("F"), column_index_from_string("F"))}': {'F'}
	}

	cf_list = {
		f"ISBLANK(B3)": {'B'},
		f'AND(ISBLANK(W3),OR(T3="Mail",T3="XML",T3="weiterer"))': {'W'},
		f'AND(ISBLANK(U3),T3="SAP-Formular")': {'U', 'V'},
		f'ISBLANK(~3)': {'H', 'I', 'D', 'O', 'T', 'X', 'Z'}
	}

	wb.conditional_formatting = ConditionalFormattingList()
	btt.data_validations = DataValidationList()
	wb["Übersicht"].data_validations = DataValidationList()
	add_cf(btt, cf_list)
	add_dv(btt, dv_list)
	add_dv(wb["Übersicht"], {f'=Übersicht!$E$2:$E${get_max_row(wb["Übersicht"], column_index_from_string("E"), column_index_from_string("E"))}': {'A'}})

	wb.save(file_name)
	wb.close()


def main():
	btt_concat(split=False)


if __name__ == "__main__":
	warnings.simplefilter("ignore", category=UserWarning)
	tp_mapping = {
		"Finanzen": "FI",
		"Instandhaltung": "IH",
		"Beschaffung": "BLQ",
		"Hauptleistung": "HL",
		"Nebenleistungen": "NL",
		"Reporting": "Reporting",
		"Berechtigung": "Berechtigung",
		"Archivierung": "Archivierung",
		"Stammdaten": "Stammdaten",
		"Master": "Master",
		"SAP Basis": "SAP Basis",
		"PS/IM": "PS/IM"
	}

	start_time = time.time()
	file_name = f"BTT_Master_konsolidiert_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

	main()
	end_time = time.time()
	print("execution time:", end_time - start_time)

#   TODO:
#	   - Zirkelbezug beim kopieren -> Ersatz?
#	   - Datenvalidierung, ConditionalFormatting
